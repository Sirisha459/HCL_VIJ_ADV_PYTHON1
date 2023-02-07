import openpyxl
wb=openpyxl.load_workbook("..//data/student.xlsx")
sheet=wb.active
data=sheet['A3'].value
data3=sheet.cell(row=2,column=3).value
sheet['A3'].value="hello"
print(data)
print(data3)
print(sheet.max_row)
print(sheet.max_column)
for i in range(2,12):
    print(sheet.cell(row=i,column=1).value)
